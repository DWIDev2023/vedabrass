import re
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from core.models import Product, FAQ


# Words that add no search/filter value as standalone keywords.
STOPWORDS = {
    "for", "and", "the", "of", "a", "an", "to", "in", "on", "with", "x",
}

# How many FAQ/answer column pairs exist in the sheet.
FAQ_COLUMN_COUNT = 10

# Starting Excel column index (0-based) for FAQ1 -> the 5th column (index 4)
FAQ_START_COL = 4

DEFAULT_CATEGORY = "Product"


def is_number_token(token: str) -> bool:
    """True for pure numeric tokens like '3', '5', '25' (used to split decimals in slugs)."""
    return bool(re.fullmatch(r"\d+", token))


def extract_keywords(slug: str, limit: int = 12) -> str:
    """
    Build a comma separated list of 1-2 word keywords from a slug.

    Strips size/number tokens (e.g. the '3', '2', '5' in '...-3-x-2-5-...')
    and filler stopwords, then returns unique single words followed by
    unique adjacent word-pairs (bigrams), capped at `limit` entries.
    """
    tokens = slug.lower().split("-")

    words = []
    for token in tokens:
        if is_number_token(token) or token in STOPWORDS or not token:
            continue
        words.append(token)

    # de-duplicate while preserving order
    seen = set()
    clean_words = []
    for word in words:
        if word not in seen:
            clean_words.append(word)
            seen.add(word)

    keywords = []
    kw_seen = set()

    for word in clean_words:
        if word not in kw_seen:
            keywords.append(word)
            kw_seen.add(word)

    for i in range(len(clean_words) - 1):
        bigram = f"{clean_words[i]} {clean_words[i + 1]}"
        if bigram not in kw_seen:
            keywords.append(bigram)
            kw_seen.add(bigram)

    return ", ".join(keywords[:limit])


class Command(BaseCommand):
    help = "Seed 10 FAQs per product from an Excel sheet, matched by product slug."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="faqs.xlsx",
            help="Path to the FAQ Excel file (default: faqs.xlsx in current dir).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report what would be created, without writing to the DB.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is required for this command. Install it with: pip install openpyxl"
            )

        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        dry_run = options["dry_run"]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise CommandError("Excel sheet is empty.")

        data_rows = rows[1:]  # skip header

        created_count = 0
        skipped_existing = 0
        missing_products = []

        for row in data_rows:
            if not row or len(row) < FAQ_START_COL:
                continue

            slug = row[3]
            if not slug:
                continue
            slug = str(slug).strip()

            try:
                product = Product.objects.get(slug=slug)
            except Product.DoesNotExist:
                missing_products.append(slug)
                continue

            keywords = extract_keywords(slug)

            faqs_to_create = []
            for i in range(FAQ_COLUMN_COUNT):
                q_idx = FAQ_START_COL + (i * 2)
                a_idx = q_idx + 1

                if a_idx >= len(row):
                    break

                question = row[q_idx]
                answer = row[a_idx]

                if not question or not answer:
                    continue

                question = str(question).strip()
                answer = str(answer).strip()

                if FAQ.objects.filter(product=product, question=question).exists():
                    skipped_existing += 1
                    continue

                faqs_to_create.append(
                    FAQ(
                        category=DEFAULT_CATEGORY,
                        question=question,
                        answer=answer,
                        keywords=keywords,
                        product=product,
                        is_active=True,
                    )
                )

            if faqs_to_create:
                if dry_run:
                    self.stdout.write(
                        f"[DRY RUN] Would create {len(faqs_to_create)} FAQs for '{slug}'"
                    )
                else:
                    with transaction.atomic():
                        FAQ.objects.bulk_create(faqs_to_create)
                created_count += len(faqs_to_create)

        self.stdout.write(self.style.SUCCESS(f"FAQs created: {created_count}"))
        if skipped_existing:
            self.stdout.write(
                self.style.WARNING(f"FAQs skipped (already existed): {skipped_existing}")
            )
        if missing_products:
            self.stdout.write(
                self.style.ERROR(
                    f"Products not found for {len(missing_products)} slug(s):"
                )
            )
            for s in missing_products:
                self.stdout.write(f"  - {s}")

# Run Command: python manage.py seed_product_faqs --file faqs.xlsx