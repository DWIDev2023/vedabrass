"""
Seed meta_title / meta_description for Category and Collection from an
Excel sheet, matched by slug.

Excel format expected (columns, in order):
    #, Page, URL, slug, Meta Title, Meta Description

Matching logic (per row):
    1. Try Category.objects.filter(slug=slug).
    2. If no Category match, try Collection.objects.filter(slug=slug).
    3. If a slug matches more than one Collection (Collection.slug is not
       unique in the model, unlike Category.slug), skip it and report it
       as ambiguous rather than guessing which one to update.
    4. If a slug matches neither model, report it as unmatched — this is
       expected for rows that represent filtered/listing pages rather
       than an actual Category or Collection record (e.g. 'new-arrivals',
       festival/gifting landing pages that may live on a different model
       or don't exist yet).

Requirements:
    pip install openpyxl
"""

from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from core.models import Category, Collection


class Command(BaseCommand):
    help = "Seed Category/Collection meta_title and meta_description from an Excel sheet, matched by slug."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default="metas.xlsx",
            help="Path to the metas Excel file (default: metas.xlsx in current dir).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report what would be updated, without writing to the DB.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is required for this command. Install it with: pip install openpyxl"
            )

        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        dry_run = options["dry_run"]

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise CommandError("Excel sheet is empty.")

        data_rows = rows[1:]  # skip header

        category_updates = 0
        collection_updates = 0
        ambiguous = []
        unmatched = []

        for row in data_rows:
            if not row or len(row) < 6:
                continue

            page, url, slug, meta_title, meta_description = (
                row[1], row[2], row[3], row[4], row[5]
            )
            if not slug:
                continue
            slug = str(slug).strip()
            meta_title = str(meta_title).strip() if meta_title else None
            meta_description = str(meta_description).strip() if meta_description else None

            category_qs = Category.objects.filter(slug=slug)
            if category_qs.exists():
                category = category_qs.first()
                if dry_run:
                    self.stdout.write(f"[DRY RUN] Category '{slug}' -> would update meta")
                else:
                    category.meta_title = meta_title
                    category.meta_description = meta_description
                    category.save(update_fields=["meta_title", "meta_description"])
                category_updates += 1
                continue

            collection_qs = Collection.objects.filter(slug=slug)
            count = collection_qs.count()
            if count == 1:
                collection = collection_qs.first()
                if dry_run:
                    self.stdout.write(f"[DRY RUN] Collection '{slug}' -> would update meta")
                else:
                    collection.meta_title = meta_title
                    collection.meta_description = meta_description
                    collection.save(update_fields=["meta_title", "meta_description"])
                collection_updates += 1
                continue
            elif count > 1:
                ambiguous.append((slug, page, count))
                continue

            unmatched.append((slug, page, url))

        self.stdout.write(self.style.SUCCESS(f"Category meta updated: {category_updates}"))
        self.stdout.write(self.style.SUCCESS(f"Collection meta updated: {collection_updates}"))

        if ambiguous:
            self.stdout.write(
                self.style.WARNING(f"Ambiguous slugs (multiple Collection matches): {len(ambiguous)}")
            )
            for slug, page, count in ambiguous:
                self.stdout.write(f"  - '{slug}' ({page}) matched {count} Collections")

        if unmatched:
            self.stdout.write(
                self.style.ERROR(f"Unmatched slugs (no Category or Collection found): {len(unmatched)}")
            )
            for slug, page, url in unmatched:
                self.stdout.write(f"  - '{slug}' ({page}) -> {url}")

# Run Command: python manage.py seed_models_metas --file metas.xlsx